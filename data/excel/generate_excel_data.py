"""
Generate realistic manufacturing demo Excel data files.
Creates work_orders.xlsx, alarm_history.xlsx, and spare_parts_inventory.xlsx
with interconnected equipment IDs and realistic manufacturing data.
"""

import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

OUTPUT_DIR = "/Users/rajeshthokala/SLM/Manufacturing/.claude/worktrees/loving-fermi-61ff19/codebase/hybrid-graphrag-manufacturing/data/excel/"

# ──────────────────────────────────────────────────────────
# Shared reference data
# ──────────────────────────────────────────────────────────

EQUIPMENT = {
    "P-201": "Centrifugal Pump A",
    "P-202": "Centrifugal Pump B",
    "P-203": "Positive Displacement Pump",
    "CV-301": "Belt Conveyor Main",
    "CV-302": "Belt Conveyor Secondary",
    "CV-303": "Screw Conveyor",
    "HP-401": "Hydraulic Press Primary",
    "HP-402": "Hydraulic Press Secondary",
}

TECHNICIANS = [
    "James Mitchell", "Sarah Chen", "Carlos Rivera", "Emily Watson",
    "David Okonkwo", "Maria Santos", "Robert Kim", "Angela Fischer",
    "Michael Torres", "Lisa Johansson",
]

# Alarm codes grouped by equipment type
PUMP_ALARMS = {
    "ALM-P001": "High vibration on drive-end bearing",
    "ALM-P002": "Seal leakage detected",
    "ALM-P003": "Discharge pressure below setpoint",
    "ALM-P004": "Motor overcurrent trip",
    "ALM-P005": "Bearing temperature high",
    "ALM-P006": "Cavitation detected",
    "ALM-P007": "Flow rate deviation",
    "ALM-P008": "Coupling misalignment warning",
}

CONVEYOR_ALARMS = {
    "ALM-C001": "Belt tracking deviation",
    "ALM-C002": "Belt slip detected",
    "ALM-C003": "Roller bearing failure",
    "ALM-C004": "Overload condition",
    "ALM-C005": "Emergency stop activated",
    "ALM-C006": "Speed deviation from setpoint",
}

HYDRAULIC_ALARMS = {
    "ALM-H001": "Hydraulic oil temperature high",
    "ALM-H002": "System pressure drop",
    "ALM-H003": "Oil contamination warning",
    "ALM-H004": "Cylinder seal leak",
    "ALM-H005": "Accumulator pre-charge low",
    "ALM-H006": "Filter differential pressure high",
}

PUMP_EQUIPMENT = ["P-201", "P-202", "P-203"]
CONVEYOR_EQUIPMENT = ["CV-301", "CV-302", "CV-303"]
HYDRAULIC_EQUIPMENT = ["HP-401", "HP-402"]


def get_alarms_for_equipment(eq_id):
    if eq_id.startswith("P-"):
        return PUMP_ALARMS
    elif eq_id.startswith("CV-"):
        return CONVEYOR_ALARMS
    else:
        return HYDRAULIC_ALARMS


def style_header(ws, num_cols):
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)


# ──────────────────────────────────────────────────────────
# 1. work_orders.xlsx  (50 records)
# ──────────────────────────────────────────────────────────

SYMPTOM_MAP = {
    "ALM-P001": [
        "High vibration detected on bearing DE side, amplitude 12.5 mm/s RMS",
        "Intermittent vibration spike on drive-end bearing exceeding 10 mm/s",
        "Abnormal vibration pattern detected during startup sequence",
    ],
    "ALM-P002": [
        "Mechanical seal leakage rate exceeding 5 drops/min",
        "Visible fluid weeping from primary seal face",
        "Secondary containment sump level rising due to seal leak",
    ],
    "ALM-P003": [
        "Discharge pressure dropped to 2.1 bar, setpoint 3.5 bar",
        "Unable to maintain discharge pressure under normal load",
        "Gradual pressure decay observed over 8-hour shift",
    ],
    "ALM-P004": [
        "Motor tripping on overcurrent at 125% FLA during startup",
        "Repeated overcurrent faults during high-load operation",
        "Motor drawing 15A above rated current under normal conditions",
    ],
    "ALM-P005": [
        "Bearing temperature reached 95 deg C, alarm setpoint 85 deg C",
        "Non-drive end bearing running 20 deg C above baseline",
        "Rapid temperature rise on DE bearing after maintenance",
    ],
    "ALM-P006": [
        "Audible cavitation noise from pump suction side",
        "NPSH margin reduced to 0.3m, minimum required 1.2m",
        "Suction pressure fluctuation causing intermittent cavitation",
    ],
    "ALM-P007": [
        "Flow rate dropped 30% below design capacity",
        "Erratic flow readings on discharge flowmeter",
        "Flow rate inconsistent between shifts, suspected impeller wear",
    ],
    "ALM-P008": [
        "Laser alignment check shows 0.15mm offset, tolerance 0.05mm",
        "Coupling rubber element showing signs of accelerated wear",
        "Vibration spectrum indicates coupling misalignment at 1x RPM",
    ],
    "ALM-C001": [
        "Belt tracking deviation exceeding 15mm on return side",
        "Belt drifting to left side consistently under load",
        "Tracking roller adjustment unable to correct belt wander",
    ],
    "ALM-C002": [
        "Belt slippage detected at head pulley, speed loss 8%",
        "Drive drum surface worn, insufficient friction coefficient",
        "Belt tension below minimum causing intermittent slip at startup",
    ],
    "ALM-C003": [
        "Carrying roller seized on section 3, belt damage risk",
        "Multiple idler rollers showing bearing noise in zone 2",
        "Return roller bearing failure causing belt edge damage",
    ],
    "ALM-C004": [
        "Conveyor load exceeding 110% rated capacity",
        "Motor current spiking during product surge from upstream",
        "Overload trip activated during shift changeover material surge",
    ],
    "ALM-C005": [
        "Emergency stop pulled at station 4 due to material spillage",
        "E-stop activated during maintenance lockout procedure",
        "Safety pull-cord triggered on south side of conveyor run",
    ],
    "ALM-C006": [
        "Belt speed 12% below setpoint under normal loading",
        "Variable frequency drive not reaching commanded speed",
        "Speed fluctuation +/- 5% observed during steady-state operation",
    ],
    "ALM-H001": [
        "Hydraulic oil temperature reached 72 deg C, limit 65 deg C",
        "Oil cooler performance degraded, delta-T across cooler only 3 deg C",
        "System temperature rising steadily during continuous press cycles",
    ],
    "ALM-H002": [
        "System pressure dropped from 210 bar to 185 bar during press cycle",
        "Unable to maintain working pressure during rapid advance",
        "Pressure decay rate 5 bar/min with valves closed",
    ],
    "ALM-H003": [
        "Particle count ISO 21/19/16, target 18/16/13",
        "Water contamination detected in oil sample, 850 ppm",
        "Oil analysis shows elevated copper and iron particle levels",
    ],
    "ALM-H004": [
        "Main cylinder rod seal leaking, visible oil on rod surface",
        "Piston seal bypass causing slow drift under load hold",
        "Cylinder extending speed reduced 40% due to internal bypass",
    ],
    "ALM-H005": [
        "Bladder accumulator pre-charge at 60 bar, spec 90 bar",
        "Accumulator unable to dampen pressure pulsations effectively",
        "Response time degraded due to low accumulator pre-charge",
    ],
    "ALM-H006": [
        "Return line filter dP at 3.2 bar, bypass set at 3.5 bar",
        "Pressure filter element near collapse pressure",
        "Filter indicator showing bypass imminent, element saturated",
    ],
}

ROOT_CAUSE_MAP = {
    "ALM-P001": [
        "Bearing inner race spalling due to fatigue after 18,000 hours",
        "Rotor imbalance from impeller erosion",
        "Foundation bolt loosening causing resonance at operating speed",
    ],
    "ALM-P002": [
        "Mechanical seal carbon face worn beyond tolerance",
        "Seal spring tension lost due to chemical attack on elastomers",
        "Shaft sleeve scoring causing seal face misalignment",
    ],
    "ALM-P003": [
        "Impeller vane erosion reducing pump efficiency by 15%",
        "Wear ring clearance doubled from design specification",
        "Suction strainer 60% blocked with debris",
    ],
    "ALM-P004": [
        "Motor winding insulation degradation on phase B",
        "VFD power module partial failure causing current imbalance",
        "Impeller rubbing on casing due to shaft deflection",
    ],
    "ALM-P005": [
        "Bearing lubrication grease degraded, insufficient oil film",
        "Bearing cage wear from contaminated lubricant",
        "Excessive preload on angular contact bearing pair",
    ],
    "ALM-P006": [
        "Suction line partially blocked by foreign object",
        "Tank level dropped below minimum NPSH requirement",
        "Suction valve inadvertently left partially closed",
    ],
    "ALM-P007": [
        "Impeller wear reducing head capacity by 25%",
        "Check valve on discharge partially stuck closed",
        "Air entrainment from tank vortex at low level",
    ],
    "ALM-P008": [
        "Thermal growth compensation not applied during cold alignment",
        "Motor soft foot condition, 0.12mm gap on rear foot",
        "Baseplate grout deterioration allowing frame movement",
    ],
    "ALM-C001": [
        "Conveyor structure out of level by 8mm over 20m span",
        "Material buildup on tail pulley causing uneven belt tension",
        "Idler roller seized on one side, creating lateral force",
    ],
    "ALM-C002": [
        "Head pulley lagging worn smooth, reduced friction coefficient",
        "Belt tension spring take-up reached end of travel",
        "Wet operating conditions reducing belt-to-pulley friction",
    ],
    "ALM-C003": [
        "Roller bearing seal failed allowing moisture ingress",
        "Bearing grease fitting blocked, no lubrication for 6 months",
        "Manufacturing defect in bearing inner race",
    ],
    "ALM-C004": [
        "Upstream process upset causing material surge 3x normal rate",
        "Level sensor on feed hopper failed in full position",
        "Operator error, manual feed rate set too high",
    ],
    "ALM-C005": [
        "Material spillage from belt edge due to overloading",
        "Personnel safety concern during nearby maintenance activity",
        "False activation from damaged pull-cord switch",
    ],
    "ALM-C006": [
        "VFD encoder feedback cable intermittent connection",
        "Drive belt between motor and gearbox slipping",
        "Gearbox output bearing wear causing speed loss",
    ],
    "ALM-H001": [
        "Oil cooler fins fouled with airborne dust and debris",
        "Cooling fan motor failed, running on natural convection only",
        "Thermostat bypass valve stuck open, bypassing cooler",
    ],
    "ALM-H002": [
        "Directional control valve spool wear allowing internal leakage",
        "Pump compensator spring fatigued, reducing output pressure",
        "High-pressure hose fitting weeping at crimp connection",
    ],
    "ALM-H003": [
        "Reservoir breather cap damaged allowing dust ingress",
        "Heat exchanger tube leak introducing water to oil system",
        "Wear metals from pump gear set approaching end of life",
    ],
    "ALM-H004": [
        "Rod seal hardened from operating at elevated temperatures",
        "Cylinder bore scored from contaminated oil particle damage",
        "Seal extrusion from pressure spikes exceeding 250 bar",
    ],
    "ALM-H005": [
        "Bladder permeation over time, nitrogen loss rate 2 bar/month",
        "Bladder puncture from contamination particle impact",
        "Incorrect pre-charge gas used during last service",
    ],
    "ALM-H006": [
        "Normal filter element life exceeded, 2,100 hours vs 2,000 spec",
        "Process contamination event introduced abnormal particulate",
        "Incorrect filter micron rating installed during last change",
    ],
}

RESOLUTION_MAP = {
    "ALM-P001": [
        "Replaced DE bearing (SKF 6310-2RS), balanced rotor, verified vibration < 4 mm/s",
        "Installed new bearing set, re-aligned coupling, torqued foundation bolts to spec",
        "Replaced both bearings, applied proper grease fill (40% cavity), test run satisfactory",
    ],
    "ALM-P002": [
        "Replaced mechanical seal assembly (John Crane Type 21), pressure tested",
        "Installed new seal, replaced shaft sleeve, verified zero leakage at operating pressure",
        "Replaced seal faces and elastomers, flushed seal plan piping, commissioned seal support system",
    ],
    "ALM-P003": [
        "Replaced impeller and wear rings, restored pump to design performance curve",
        "Cleaned suction strainer, replaced wear rings, verified discharge pressure at 3.6 bar",
        "Installed new impeller, adjusted clearances per OEM spec, recorded new baseline data",
    ],
    "ALM-P004": [
        "Rewound motor stator, replaced VFD IGBTs, load tested to 110% FLA",
        "Replaced motor with refurbished spare, sent original for rewind",
        "Replaced VFD power module, updated motor parameters, verified current balance within 2%",
    ],
    "ALM-P005": [
        "Repacked bearings with Mobilith SHC 100, adjusted preload per procedure",
        "Replaced bearing set, flushed housing, applied correct grease quantity",
        "Replaced bearings, installed proximity probe for continuous temperature monitoring",
    ],
    "ALM-P006": [
        "Removed foreign object from suction line, inspected strainer integrity",
        "Adjusted tank level low-low interlock setpoint, verified NPSH margin > 2m",
        "Replaced suction valve actuator, confirmed full-open position with limit switch",
    ],
    "ALM-P007": [
        "Replaced impeller and volute casing, restored pump to nameplate capacity",
        "Freed stuck check valve, replaced disc and seat, tested for full closure",
        "Installed vortex breaker in tank, adjusted minimum operating level",
    ],
    "ALM-P008": [
        "Performed hot alignment with thermal targets, achieved < 0.03mm offset",
        "Corrected soft foot with precision shims, re-aligned within OEM tolerance",
        "Repaired baseplate grout, re-leveled assembly, performed precision laser alignment",
    ],
    "ALM-C001": [
        "Re-leveled conveyor structure, adjusted training idlers, verified tracking under load",
        "Cleaned tail pulley, installed belt scraper, adjusted tracking idlers",
        "Replaced seized idler roller, re-tensioned belt, tracking within 5mm tolerance",
    ],
    "ALM-C002": [
        "Re-lagged head pulley with ceramic lagging, restored friction coefficient",
        "Adjusted gravity take-up counterweight, replaced worn lagging on drive pulley",
        "Installed new drive pulley lagging, adjusted belt tension to 2.5% elongation",
    ],
    "ALM-C003": [
        "Replaced failed rollers in section 3, greased all rollers per schedule",
        "Replaced 12 idler rollers in zone 2, implemented quarterly lubrication route",
        "Replaced return roller assembly, inspected and cleaned adjacent rollers",
    ],
    "ALM-C004": [
        "Calibrated upstream level sensor, adjusted feed rate control loop",
        "Installed overload protection relay, calibrated belt scale, adjusted material flow",
        "Retrained operators on manual feed rate procedures, added visual feed rate indicator",
    ],
    "ALM-C005": [
        "Cleared spillage, adjusted belt loading to prevent recurrence, reset E-stop",
        "Released E-stop after maintenance activity completed and area cleared",
        "Replaced damaged pull-cord switch, tested all E-stop stations on conveyor run",
    ],
    "ALM-C006": [
        "Replaced encoder cable and connectors, verified speed feedback accuracy",
        "Replaced motor-to-gearbox V-belts, tensioned to manufacturer spec",
        "Replaced gearbox output bearing, refilled with synthetic gear oil, tested speed accuracy",
    ],
    "ALM-H001": [
        "Cleaned oil cooler fins with compressed air, replaced cooling fan motor",
        "Replaced cooling fan motor, cleaned cooler core, verified oil temp < 55 deg C",
        "Replaced thermostat bypass valve, cleaned cooler fins, oil temperature normalized",
    ],
    "ALM-H002": [
        "Replaced directional control valve spool assembly, verified zero internal leakage",
        "Replaced pump compensator spring, adjusted deadband, verified 210 bar output",
        "Re-crimped high-pressure hose fitting, pressure tested to 315 bar for 5 minutes",
    ],
    "ALM-H003": [
        "Replaced reservoir breather with desiccant type, changed oil and filters",
        "Repaired heat exchanger tube leak, performed oil flush and kidney-loop filtration",
        "Replaced pump gear set, flushed system, new oil charge, verified particle count ISO 17/15/12",
    ],
    "ALM-H004": [
        "Replaced rod and piston seals, honed cylinder bore, pressure tested to 250 bar",
        "Replaced complete seal kit, polished rod surface, verified zero drift under load",
        "Installed new rod seal with backup ring, verified system pressure holding for 30 min",
    ],
    "ALM-H005": [
        "Recharged accumulator bladder with nitrogen to 90 bar, checked for leaks",
        "Replaced bladder assembly, pre-charged to spec, verified response time",
        "Recharged with correct nitrogen specification, installed pre-charge check gauge",
    ],
    "ALM-H006": [
        "Replaced filter element with correct 10-micron rating, reset dP indicator",
        "Installed new filter element, flushed system, verified dP < 0.5 bar clean",
        "Replaced filter element and housing O-ring, verified no bypass flow",
    ],
}

PRIORITIES = ["Critical", "High", "Medium", "Low"]
PRIORITY_WEIGHTS = [0.1, 0.25, 0.45, 0.2]


def generate_work_orders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Orders"

    headers = [
        "work_order_id", "equipment_id", "alarm_code", "symptom_description",
        "root_cause", "resolution", "downtime_hours", "cost_usd",
        "priority", "date_reported", "date_resolved", "technician",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    all_equipment = list(EQUIPMENT.keys())

    for i in range(1, 51):
        wo_id = f"WO-2024-{i:03d}"
        eq_id = random.choice(all_equipment)
        alarms = get_alarms_for_equipment(eq_id)
        alarm_code = random.choice(list(alarms.keys()))

        symptom = random.choice(SYMPTOM_MAP[alarm_code])
        root_cause = random.choice(ROOT_CAUSE_MAP[alarm_code])
        resolution = random.choice(RESOLUTION_MAP[alarm_code])

        priority = random.choices(PRIORITIES, weights=PRIORITY_WEIGHTS, k=1)[0]

        # Downtime and cost correlated with priority
        if priority == "Critical":
            downtime = round(random.uniform(12, 48), 1)
            cost = round(random.uniform(8000, 25000), 2)
        elif priority == "High":
            downtime = round(random.uniform(4, 24), 1)
            cost = round(random.uniform(3000, 15000), 2)
        elif priority == "Medium":
            downtime = round(random.uniform(1, 8), 1)
            cost = round(random.uniform(500, 5000), 2)
        else:
            downtime = round(random.uniform(0.5, 3), 1)
            cost = round(random.uniform(100, 1500), 2)

        date_reported = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 364))
        resolution_delay_hours = downtime + random.uniform(0, 72)
        date_resolved = date_reported + timedelta(hours=resolution_delay_hours)

        technician = random.choice(TECHNICIANS)

        ws.append([
            wo_id, eq_id, alarm_code, symptom, root_cause, resolution,
            downtime, cost, priority,
            date_reported.strftime("%Y-%m-%d"),
            date_resolved.strftime("%Y-%m-%d %H:%M"),
            technician,
        ])

    auto_width(ws)
    filepath = OUTPUT_DIR + "work_orders.xlsx"
    wb.save(filepath)
    print(f"Created {filepath} with 50 work order records")


# ──────────────────────────────────────────────────────────
# 2. alarm_history.xlsx  (200 records)
# ──────────────────────────────────────────────────────────

SEVERITY_BY_ALARM = {
    "ALM-P001": "Warning", "ALM-P002": "Warning", "ALM-P003": "Warning",
    "ALM-P004": "Critical", "ALM-P005": "Warning", "ALM-P006": "Warning",
    "ALM-P007": "Info", "ALM-P008": "Info",
    "ALM-C001": "Info", "ALM-C002": "Warning", "ALM-C003": "Warning",
    "ALM-C004": "Critical", "ALM-C005": "Critical", "ALM-C006": "Warning",
    "ALM-H001": "Warning", "ALM-H002": "Critical", "ALM-H003": "Info",
    "ALM-H004": "Warning", "ALM-H005": "Warning", "ALM-H006": "Info",
}

ACTION_TAKEN_MAP = {
    "Critical": [
        "Immediate shutdown and emergency maintenance initiated",
        "Equipment isolated, maintenance team dispatched, backup activated",
        "Production halted, root cause investigation started immediately",
        "Emergency repair crew mobilized, spare parts pulled from inventory",
    ],
    "Warning": [
        "Condition monitored, maintenance scheduled for next planned outage",
        "Operating parameters adjusted to reduce severity, PM scheduled",
        "Alarm acknowledged, work order raised for corrective maintenance",
        "Temporary operational adjustment applied, permanent fix planned",
        "Increased monitoring frequency, parts ordered for scheduled repair",
    ],
    "Info": [
        "Logged for trend analysis, no immediate action required",
        "Noted in shift handover log, to be reviewed by reliability engineer",
        "Added to condition monitoring report for weekly review",
        "Acknowledged, parameter within acceptable range but trending",
    ],
}


def generate_alarm_history():
    wb = Workbook()
    ws = wb.active
    ws.title = "Alarm History"

    headers = [
        "timestamp", "equipment_id", "alarm_code", "alarm_description",
        "severity", "duration_minutes", "acknowledged_by", "action_taken",
        "recurring",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    all_alarms = {**PUMP_ALARMS, **CONVEYOR_ALARMS, **HYDRAULIC_ALARMS}
    records = []

    # Track recurring alarms: some equipment+alarm combos appear multiple times
    recurring_combos = set()
    combo_counts = {}

    for _ in range(200):
        eq_id = random.choice(list(EQUIPMENT.keys()))
        alarms = get_alarms_for_equipment(eq_id)
        alarm_code = random.choice(list(alarms.keys()))
        alarm_desc = alarms[alarm_code]

        combo_key = (eq_id, alarm_code)
        combo_counts[combo_key] = combo_counts.get(combo_key, 0) + 1
        if combo_counts[combo_key] > 1:
            recurring_combos.add(combo_key)

        severity = SEVERITY_BY_ALARM[alarm_code]
        # Sometimes override severity based on context
        if random.random() < 0.15:
            severity = random.choice(["Critical", "Warning"])

        if severity == "Critical":
            duration = random.randint(30, 480)
        elif severity == "Warning":
            duration = random.randint(10, 240)
        else:
            duration = random.randint(1, 60)

        ts = datetime(2024, 1, 1) + timedelta(
            days=random.randint(0, 364),
            hours=random.randint(0, 23),
            minutes=random.randint(0, 59),
        )

        action = random.choice(ACTION_TAKEN_MAP[severity])
        tech = random.choice(TECHNICIANS)

        records.append((ts, eq_id, alarm_code, alarm_desc, severity,
                         duration, tech, action, combo_key))

    # Sort by timestamp
    records.sort(key=lambda r: r[0])

    for rec in records:
        ts, eq_id, alarm_code, alarm_desc, severity, duration, tech, action, combo_key = rec
        recurring = "Yes" if combo_key in recurring_combos else "No"
        ws.append([
            ts.strftime("%Y-%m-%d %H:%M:%S"),
            eq_id, alarm_code, alarm_desc, severity,
            duration, tech, action, recurring,
        ])

    auto_width(ws)
    filepath = OUTPUT_DIR + "alarm_history.xlsx"
    wb.save(filepath)
    print(f"Created {filepath} with 200 alarm history records")


# ──────────────────────────────────────────────────────────
# 3. spare_parts_inventory.xlsx  (40 parts)
# ──────────────────────────────────────────────────────────

SPARE_PARTS = [
    # Mechanical parts for pumps
    ("SP-1001", "SKF 6310-2RS Deep Groove Ball Bearing", "P-201, P-202", "Mechanical", 12, 4, 85.50, 5, "SKF Authorized Distributor"),
    ("SP-1002", "SKF 7310 Angular Contact Bearing", "P-201, P-202, P-203", "Mechanical", 8, 3, 142.00, 7, "SKF Authorized Distributor"),
    ("SP-1003", "John Crane Type 21 Mechanical Seal 50mm", "P-201, P-202", "Mechanical", 4, 2, 1250.00, 14, "John Crane Service Center"),
    ("SP-1004", "John Crane Type 21 Mechanical Seal 65mm", "P-203", "Mechanical", 3, 2, 1480.00, 14, "John Crane Service Center"),
    ("SP-1005", "Pump Impeller 316SS 200mm", "P-201, P-202", "Mechanical", 2, 1, 2800.00, 28, "Pump OEM Direct"),
    ("SP-1006", "Pump Wear Ring Set 316SS", "P-201, P-202, P-203", "Mechanical", 6, 3, 340.00, 10, "Pump OEM Direct"),
    ("SP-1007", "Coupling Element (Rexnord Omega E40)", "P-201, P-202", "Mechanical", 4, 2, 195.00, 7, "Rexnord Corporation"),
    ("SP-1008", "Shaft Sleeve 316SS 55mm x 120mm", "P-201, P-202", "Mechanical", 3, 2, 420.00, 14, "Pump OEM Direct"),
    # Electrical parts for pumps
    ("SP-1009", "Motor Bearing 6314-2RS (DE)", "P-201, P-202, P-203", "Electrical", 6, 3, 110.00, 5, "SKF Authorized Distributor"),
    ("SP-1010", "VFD IGBT Power Module 75kW", "P-201, P-202", "Electrical", 2, 1, 3200.00, 21, "ABB Drives Service"),
    ("SP-1011", "Motor Stator Winding Assembly 75kW", "P-201, P-202", "Electrical", 1, 1, 4500.00, 35, "Motor Rewind Specialists"),
    ("SP-1012", "Proximity Vibration Probe 8mm", "P-201, P-202, P-203", "Instrumentation", 8, 4, 285.00, 7, "Bently Nevada"),
    # Conveyor parts
    ("SP-2001", "Conveyor Belt Splice Kit 800mm", "CV-301, CV-302", "Mechanical", 3, 2, 650.00, 10, "Continental Belt Supply"),
    ("SP-2002", "Ceramic Pulley Lagging Kit 500mm Dia", "CV-301, CV-302", "Mechanical", 2, 1, 1850.00, 21, "Elastotec Lagging Systems"),
    ("SP-2003", "Carrying Idler Roller 133mm x 600mm", "CV-301, CV-302", "Mechanical", 24, 10, 45.00, 5, "Precision Roller Company"),
    ("SP-2004", "Return Idler Roller 102mm x 800mm", "CV-301, CV-302", "Mechanical", 18, 8, 38.00, 5, "Precision Roller Company"),
    ("SP-2005", "Training Idler Assembly (Self-aligning)", "CV-301, CV-302", "Mechanical", 6, 3, 320.00, 10, "Martin Engineering"),
    ("SP-2006", "Belt Scraper Blade Primary 800mm", "CV-301, CV-302", "Mechanical", 8, 4, 125.00, 5, "Martin Engineering"),
    ("SP-2007", "Take-Up Screw Assembly M24 x 500mm", "CV-301, CV-302", "Mechanical", 4, 2, 180.00, 7, "Conveyor Components Co"),
    ("SP-2008", "Screw Conveyor Flight Section 300mm", "CV-303", "Mechanical", 4, 2, 520.00, 14, "Screw Conveyor Corp"),
    ("SP-2009", "Screw Conveyor Hanger Bearing", "CV-303", "Mechanical", 6, 3, 165.00, 10, "Screw Conveyor Corp"),
    ("SP-2010", "Belt Speed Sensor (Encoder)", "CV-301, CV-302", "Instrumentation", 3, 2, 390.00, 7, "Pepperl+Fuchs"),
    ("SP-2011", "Pull-Cord Safety Switch", "CV-301, CV-302", "Electrical", 4, 2, 210.00, 5, "Allen-Bradley"),
    ("SP-2012", "Conveyor Motor V-Belt Set (5x SPA)", "CV-301, CV-302", "Mechanical", 6, 3, 95.00, 3, "Gates Industrial"),
    # Hydraulic press parts
    ("SP-3001", "Hydraulic Cylinder Rod Seal Kit 100mm", "HP-401, HP-402", "Hydraulic", 6, 3, 385.00, 10, "Parker Hannifin"),
    ("SP-3002", "Hydraulic Cylinder Piston Seal Kit 160mm", "HP-401, HP-402", "Hydraulic", 4, 2, 520.00, 10, "Parker Hannifin"),
    ("SP-3003", "Directional Control Valve Spool Assembly", "HP-401, HP-402", "Hydraulic", 2, 1, 2100.00, 21, "Bosch Rexroth"),
    ("SP-3004", "High Pressure Filter Element 10 micron", "HP-401, HP-402", "Hydraulic", 12, 6, 75.00, 5, "Hydac Technology"),
    ("SP-3005", "Return Line Filter Element 25 micron", "HP-401, HP-402", "Hydraulic", 10, 5, 55.00, 5, "Hydac Technology"),
    ("SP-3006", "Accumulator Bladder Assembly 10L", "HP-401, HP-402", "Hydraulic", 3, 2, 680.00, 14, "Parker Hannifin"),
    ("SP-3007", "Hydraulic Pump Gear Set", "HP-401, HP-402", "Hydraulic", 2, 1, 3400.00, 28, "Bosch Rexroth"),
    ("SP-3008", "Oil Cooler Fan Motor 1.5kW", "HP-401, HP-402", "Electrical", 2, 1, 480.00, 10, "Siemens Industrial"),
    ("SP-3009", "Thermostat Bypass Valve DN25", "HP-401, HP-402", "Hydraulic", 3, 2, 290.00, 7, "Danfoss"),
    ("SP-3010", "HP Hose Assembly DN12 x 2m (350 bar)", "HP-401, HP-402", "Hydraulic", 6, 3, 185.00, 5, "Parker Hannifin"),
    ("SP-3011", "Pressure Transducer 0-400 bar", "HP-401, HP-402", "Instrumentation", 4, 2, 450.00, 7, "WIKA Instruments"),
    ("SP-3012", "Hydraulic Oil ISO VG46 (20L drum)", "HP-401, HP-402", "Hydraulic", 8, 4, 120.00, 3, "Shell Lubricants"),
    # General / cross-equipment
    ("SP-4001", "Desiccant Breather Cap DN50", "HP-401, HP-402", "Hydraulic", 5, 3, 65.00, 5, "Des-Case Corporation"),
    ("SP-4002", "Mobilith SHC 100 Grease Cartridge 400g", "P-201, P-202, P-203, CV-301, CV-302", "Mechanical", 20, 10, 18.50, 3, "ExxonMobil Lubricants"),
    ("SP-4003", "Synthetic Gear Oil ISO 220 (5L)", "CV-301, CV-302, CV-303", "Mechanical", 6, 3, 85.00, 5, "Shell Lubricants"),
    ("SP-4004", "Temperature RTD Probe PT100 6mm x 100mm", "P-201, P-202, P-203, HP-401, HP-402", "Instrumentation", 10, 5, 95.00, 5, "Endress+Hauser"),
]


def generate_spare_parts():
    wb = Workbook()
    ws = wb.active
    ws.title = "Spare Parts Inventory"

    headers = [
        "part_number", "part_name", "equipment_compatibility", "category",
        "quantity_in_stock", "reorder_level", "unit_cost_usd",
        "lead_time_days", "supplier", "last_used_date",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for part in SPARE_PARTS:
        pn, name, compat, cat, qty, reorder, cost, lead, supplier = part
        last_used = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 364))
        ws.append([
            pn, name, compat, cat,
            qty, reorder, cost, lead, supplier,
            last_used.strftime("%Y-%m-%d"),
        ])

    auto_width(ws)
    filepath = OUTPUT_DIR + "spare_parts_inventory.xlsx"
    wb.save(filepath)
    print(f"Created {filepath} with {len(SPARE_PARTS)} spare parts records")


# ──────────────────────────────────────────────────────────
# Generate all files
# ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    generate_work_orders()
    generate_alarm_history()
    generate_spare_parts()
    print("\nAll Excel files generated successfully!")
