"""
Document Ingestion Layer — Unified parsers for PDF, TXT, and Excel files.
Extracts text and metadata from each format into a common Document structure.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Iterable

import openpyxl
import pdfplumber

try:
    from config import SUPPORTED_EXTENSIONS
except ImportError:
    from .config import SUPPORTED_EXTENSIONS


logger = logging.getLogger("doc_pipeline.ingestion")


class DocType(str, Enum):
    PDF = "pdf"
    TXT = "txt"
    EXCEL = "excel"


@dataclass
class Document:
    content: str
    metadata: dict = field(default_factory=dict)
    source: str = ""
    doc_type: str = ""


class PDFParser:
    """Extracts text from PDF files using pdfplumber, preserving table structure."""

    def parse(self, file_path: str | Path) -> list[Document]:
        path = Path(file_path)
        documents: list[Document] = []

        with pdfplumber.open(str(path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text_parts: list[str] = []

                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    text_parts.append(self._format_table(table))

                page_text = page.extract_text() or ""
                if page_text.strip():
                    text_parts.insert(0, page_text)

                full_text = "\n\n".join(text_parts).strip()
                if not full_text:
                    continue

                documents.append(Document(
                    content=full_text,
                    metadata={
                        "page": page_num,
                        "total_pages": len(pdf.pages),
                        "filename": path.name,
                        "has_tables": len(tables) > 0,
                    },
                    source=str(path),
                    doc_type=DocType.PDF.value,
                ))

        return documents

    @staticmethod
    def _format_table(table: list[list]) -> str:
        header = table[0]
        rows = table[1:]
        col_widths = [max(len(str(cell or "")) for cell in col) for col in zip(*table)]
        out = " | ".join(str(h or "") for h in header) + "\n"
        out += "-" * sum(col_widths) + "\n"
        for row in rows:
            out += " | ".join(str(cell or "") for cell in row) + "\n"
        return out


class TXTParser:
    """Parses plain text files, preserving section structure."""

    def parse(self, file_path: str | Path) -> list[Document]:
        path = Path(file_path)
        text = path.read_text(encoding="utf-8")

        sections = self._split_sections(text)
        return [
            Document(
                content=content.strip(),
                metadata={
                    "section_index": i,
                    "section_title": title,
                    "filename": path.name,
                    "total_sections": len(sections),
                },
                source=str(path),
                doc_type=DocType.TXT.value,
            )
            for i, (title, content) in enumerate(sections)
        ]

    @staticmethod
    def _split_sections(text: str) -> list[tuple[str, str]]:
        sections: list[tuple[str, str]] = []
        current_title = "Header"
        current_content: list[str] = []

        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("===") and stripped.endswith("===") and len(stripped) > 6:
                if current_content:
                    sections.append((current_title, "\n".join(current_content)))
                current_title = stripped.strip("= ")
                current_content = []
            else:
                current_content.append(line)

        if current_content:
            sections.append((current_title, "\n".join(current_content)))

        return sections


class ExcelParser:
    """Parses Excel files, converting each sheet into structured text."""

    def parse(self, file_path: str | Path) -> list[Document]:
        path = Path(file_path)
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        documents: list[Document] = []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                headers = [str(h or "") for h in rows[0]]
                data_rows = rows[1:]

                lines: list[str] = [
                    f"Sheet: {sheet_name}",
                    f"Columns: {', '.join(headers)}",
                    "",
                    f"Summary: {self._summarize(headers, data_rows)}",
                    "",
                ]
                for row in data_rows:
                    lines.append(" | ".join(
                        f"{headers[i]}: {self._format_cell(cell)}"
                        for i, cell in enumerate(row) if i < len(headers)
                    ))

                documents.append(Document(
                    content="\n".join(lines),
                    metadata={
                        "sheet_name": sheet_name,
                        "row_count": len(data_rows),
                        "columns": headers,
                        "filename": path.name,
                    },
                    source=str(path),
                    doc_type=DocType.EXCEL.value,
                ))
        finally:
            wb.close()

        return documents

    @staticmethod
    def _format_cell(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return f"{value:.4f}" if abs(value) < 1 else f"{value:.2f}"
        return str(value)

    @staticmethod
    def _summarize(headers: list[str], rows: list) -> str:
        numeric_cols: dict[str, dict[str, float]] = {}
        for i, h in enumerate(headers):
            vals = [row[i] for row in rows if i < len(row) and isinstance(row[i], (int, float))]
            if vals:
                numeric_cols[h] = {
                    "min": min(vals), "max": max(vals),
                    "avg": sum(vals) / len(vals), "count": len(vals),
                }

        parts = [f"{len(rows)} records"]
        for col, stats in list(numeric_cols.items())[:3]:
            parts.append(f"{col}: avg={stats['avg']:.1f}, range=[{stats['min']}, {stats['max']}]")
        return "; ".join(parts)


class DocumentIngestion:
    """Auto-routes a file to the appropriate parser based on its extension."""

    def __init__(self):
        self.parsers = {
            ".pdf": PDFParser(),
            ".txt": TXTParser(),
            ".xlsx": ExcelParser(),
            ".xls": ExcelParser(),
        }

    def ingest_file(self, file_path: str | Path) -> list[Document]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = path.suffix.lower()
        parser = self.parsers.get(ext)
        if not parser:
            raise ValueError(
                f"Unsupported file type: {ext}. Supported: {SUPPORTED_EXTENSIONS}"
            )

        docs = parser.parse(path)
        logger.info("Ingested %s: %d document segments", path.name, len(docs))
        return docs

    def ingest_directory(self, dir_path: str | Path) -> list[Document]:
        directory = Path(dir_path)
        if not directory.is_dir():
            raise NotADirectoryError(f"Not a directory: {dir_path}")

        supported = set(self.parsers.keys())
        all_docs: list[Document] = []

        # Recurse so classification subfolders (``management/``,
        # ``restricted/``) ship their documents through the same parser
        # pipeline. ``classify_from_path`` in ``core.document_acl`` reads
        # the folder name back out at chunking time to attach the
        # classification tag.
        for file_path in sorted(directory.rglob("*")):
            if file_path.is_file() and file_path.suffix.lower() in supported:
                all_docs.extend(self.ingest_file(file_path))

        logger.info("Total: %d document segments from %s", len(all_docs), directory)
        return all_docs
