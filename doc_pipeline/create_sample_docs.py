"""
Generate sample manufacturing business documents: PDFs, TXT, and Excel files.
These serve as realistic input data for the document processing pipeline.
"""

import os
from pathlib import Path
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).parent / "input_docs"
BASE_DIR.mkdir(exist_ok=True)

styles = getSampleStyleSheet()
title_style = ParagraphStyle(
    "CustomTitle", parent=styles["Title"],
    fontSize=18, textColor=HexColor("#1a237e"), spaceAfter=20
)
heading_style = ParagraphStyle(
    "CustomHeading", parent=styles["Heading2"],
    fontSize=14, textColor=HexColor("#283593"), spaceBefore=14, spaceAfter=8
)
body_style = ParagraphStyle(
    "CustomBody", parent=styles["Normal"],
    fontSize=10, leading=14, spaceAfter=6
)


def build_table(data, col_widths=None):
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#283593")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, HexColor("#e8eaf6")]),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
    ]))
    return t


# ── PDF 1: Quality Control Manual ──────────────────────────────────────────
def create_quality_control_pdf():
    path = str(BASE_DIR / "quality_control_manual.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    story = []

    story.append(Paragraph("Quality Control Manual — SteelForge Manufacturing", title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph("1. Purpose and Scope", heading_style))
    story.append(Paragraph(
        "This Quality Control Manual establishes the procedures, standards, and responsibilities "
        "for ensuring product quality across all SteelForge Manufacturing production lines. "
        "It covers raw material inspection, in-process quality checks, final product testing, "
        "non-conformance reporting (NCR), corrective and preventive actions (CAPA), and "
        "compliance with ISO 9001:2015 and IATF 16949 automotive quality management standards.",
        body_style
    ))
    story.append(Paragraph(
        "All production personnel, quality inspectors, and line supervisors are required to "
        "follow the procedures outlined in this manual. Deviations must be documented via the "
        "Non-Conformance Report (NCR-Form-101) and escalated within 24 hours.",
        body_style
    ))

    story.append(Paragraph("2. Incoming Material Inspection", heading_style))
    story.append(Paragraph(
        "All raw materials received at the loading dock must undergo inspection within 4 hours "
        "of delivery. The Quality Lab team uses Acceptance Quality Level (AQL) sampling per "
        "ANSI/ASQ Z1.4 standard. Critical materials such as high-tensile steel coils, aluminum "
        "billets, and copper wire stock require 100% dimensional verification using coordinate "
        "measuring machines (CMM). Chemical composition is verified via optical emission "
        "spectroscopy (OES) against material test certificates (MTC).",
        body_style
    ))
    story.append(Paragraph(
        "Supplier performance is tracked through the Supplier Scorecard system. Suppliers with "
        "a quality rating below 85% are placed on probation and require corrective action plans "
        "within 30 days. Repeat offenders (three consecutive months below threshold) are "
        "escalated to the Approved Vendor List (AVL) committee for potential removal.",
        body_style
    ))

    data = [
        ["Material", "Spec", "Test Method", "Frequency", "Accept Criteria"],
        ["Steel Coil HR", "ASTM A36", "Tensile + CMM", "Every lot", "Yield ≥ 250 MPa"],
        ["Aluminum 6061-T6", "AMS 4027", "Hardness + OES", "Every lot", "Brinell ≥ 95"],
        ["Copper Wire C110", "ASTM B3", "Conductivity", "1 per 5 lots", "≥ 100% IACS"],
        ["Stainless 304", "ASTM A240", "Corrosion test", "Every lot", "Pass 48h salt spray"],
        ["Titanium Gr5", "AMS 4928", "Ultrasonic + CMM", "100%", "No defects > 0.5mm"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data, col_widths=[1.2*inch, 1*inch, 1.2*inch, 1*inch, 1.5*inch]))

    story.append(PageBreak())
    story.append(Paragraph("3. In-Process Quality Control", heading_style))
    story.append(Paragraph(
        "Statistical Process Control (SPC) is mandatory on all CNC machining, stamping, and "
        "welding operations. Control charts (X-bar and R charts) are maintained for critical "
        "dimensions. Process capability indices (Cpk) must be ≥ 1.33 for standard features and "
        "≥ 1.67 for safety-critical dimensions. Real-time SPC data is collected via shop-floor "
        "terminals connected to the MES (Manufacturing Execution System) and monitored by "
        "Quality Engineers on the SPC Dashboard.",
        body_style
    ))
    story.append(Paragraph(
        "First Article Inspection (FAI) is required for all new part numbers, engineering change "
        "orders (ECO), and after any tooling change. FAI reports follow AS9102 format and must "
        "be approved by the Quality Manager before production release. All measurement equipment "
        "is calibrated per ISO/IEC 17025 with calibration intervals not exceeding 12 months.",
        body_style
    ))

    story.append(Paragraph("4. Final Product Testing and Release", heading_style))
    story.append(Paragraph(
        "Finished goods undergo a battery of tests before shipment release. Mechanical testing "
        "includes tensile, impact (Charpy V-notch), and fatigue testing per customer specifications. "
        "Dimensional inspection uses CMM with GD&T callouts per ASME Y14.5-2018. Surface finish "
        "is verified using profilometers (Ra values). All test results are recorded in the Enterprise "
        "Quality Management System (EQMS) and linked to the lot/batch traceability record.",
        body_style
    ))
    story.append(Paragraph(
        "Products failing final inspection are quarantined in the Material Review Board (MRB) area. "
        "The MRB team (Quality Manager, Production Manager, Engineering Lead) reviews disposition "
        "options: rework, use-as-is (with customer concession), or scrap. All MRB decisions are "
        "documented and require customer approval for safety-critical components.",
        body_style
    ))

    story.append(Paragraph("5. Non-Conformance and CAPA Process", heading_style))
    story.append(Paragraph(
        "Non-conformances are classified as Minor, Major, or Critical. Critical NCRs trigger an "
        "immediate production stop and 8D problem-solving process. Root Cause Analysis (RCA) uses "
        "5-Why, Fishbone (Ishikawa), and Fault Tree Analysis (FTA) methods. Corrective actions must "
        "be implemented within 15 business days for Critical NCRs, 30 days for Major, and 60 days "
        "for Minor. Effectiveness verification is conducted 90 days after implementation.",
        body_style
    ))

    data2 = [
        ["NCR Severity", "Response Time", "RCA Deadline", "CAPA Deadline", "Verification"],
        ["Critical", "Immediate stop", "48 hours", "15 business days", "90-day audit"],
        ["Major", "24 hours", "5 business days", "30 business days", "90-day review"],
        ["Minor", "72 hours", "10 business days", "60 business days", "Next scheduled audit"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data2))

    doc.build(story)
    print(f"  Created: {path}")


# ── PDF 2: Production Planning & Scheduling Report ─────────────────────────
def create_production_planning_pdf():
    path = str(BASE_DIR / "production_planning_report.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    story = []

    story.append(Paragraph("Production Planning & Scheduling Report — Q1 2026", title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph("1. Executive Summary", heading_style))
    story.append(Paragraph(
        "Q1 2026 production achieved 94.2% Overall Equipment Effectiveness (OEE) across Plant A "
        "and Plant B combined, exceeding the 92% target. Total output was 1,847,500 units against "
        "a planned volume of 1,900,000 units (97.2% plan attainment). The shortfall was primarily "
        "due to an unplanned 36-hour shutdown of CNC Line 4 in February caused by a spindle bearing "
        "failure. Preventive maintenance schedules have been updated to include quarterly vibration "
        "analysis on all CNC spindle assemblies.",
        body_style
    ))

    story.append(Paragraph("2. Capacity Utilization", heading_style))
    story.append(Paragraph(
        "Plant A operated at 96.1% capacity utilization with three shifts running 6 days per week. "
        "The stamping department was the bottleneck, running at 99.3% utilization. A capital request "
        "for a new 600-ton progressive stamping press (CAPEX $2.4M) has been submitted for Q3 "
        "installation to relieve this constraint. Plant B operated at 88.7% capacity, with the "
        "welding department showing the lowest utilization at 82.1% due to a skilled welder shortage. "
        "HR has initiated a recruitment campaign targeting AWS-certified welders.",
        body_style
    ))

    data = [
        ["Department", "Plant", "Capacity %", "OEE %", "Downtime Hrs", "Scrap Rate %"],
        ["CNC Machining", "A", "95.4%", "93.8%", "42", "1.2%"],
        ["Stamping", "A", "99.3%", "96.1%", "18", "0.8%"],
        ["Welding", "A", "91.2%", "89.5%", "67", "2.1%"],
        ["Assembly", "A", "94.8%", "95.2%", "31", "0.5%"],
        ["CNC Machining", "B", "92.1%", "91.3%", "55", "1.5%"],
        ["Welding", "B", "82.1%", "84.7%", "98", "2.8%"],
        ["Heat Treatment", "B", "90.5%", "92.1%", "44", "0.3%"],
        ["Finishing/Coating", "B", "88.3%", "90.6%", "52", "1.1%"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data))

    story.append(PageBreak())
    story.append(Paragraph("3. Material Requirements Planning (MRP)", heading_style))
    story.append(Paragraph(
        "MRP analysis for Q2 2026 projects a 12% increase in steel consumption driven by the "
        "new automotive transmission housing program (Part #TH-4400 series). Lead times for "
        "high-tensile steel from primary supplier Nippon Steel have increased from 8 weeks to "
        "12 weeks due to global demand. Procurement has established a secondary supply agreement "
        "with ArcelorMittal to mitigate supply risk. Safety stock levels for critical raw materials "
        "have been increased from 2 weeks to 3 weeks coverage.",
        body_style
    ))
    story.append(Paragraph(
        "The Enterprise Resource Planning (ERP) system has been configured with updated Bill of "
        "Materials (BOM) for the TH-4400 series. Routing times have been validated through time "
        "studies conducted by Industrial Engineering. The standard cost for TH-4400 is $47.82 per "
        "unit with a target cycle time of 4.2 minutes per piece on the CNC machining center.",
        body_style
    ))

    story.append(Paragraph("4. Scheduling and Sequencing", heading_style))
    story.append(Paragraph(
        "Production scheduling follows a hybrid push-pull methodology. Make-to-stock (MTS) items "
        "use weekly frozen schedules with a 4-week planning horizon. Make-to-order (MTO) items "
        "follow customer release schedules with daily sequencing. The Advanced Planning and "
        "Scheduling (APS) module optimizes job sequencing to minimize changeover times using a "
        "genetic algorithm solver. Average changeover time has been reduced from 45 minutes to "
        "28 minutes through SMED (Single-Minute Exchange of Die) initiatives.",
        body_style
    ))

    story.append(Paragraph("5. Key Performance Indicators — Q2 Targets", heading_style))
    data2 = [
        ["KPI", "Q1 Actual", "Q2 Target", "Stretch Goal"],
        ["OEE", "94.2%", "95.0%", "96.5%"],
        ["On-Time Delivery", "97.8%", "98.5%", "99.0%"],
        ["Scrap Rate", "1.3%", "< 1.0%", "< 0.8%"],
        ["Changeover Time", "28 min", "25 min", "20 min"],
        ["Inventory Turns", "8.4x", "9.0x", "10.0x"],
        ["Customer PPM", "45", "< 30", "< 20"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data2))

    doc.build(story)
    print(f"  Created: {path}")


# ── PDF 3: Safety and Compliance Report ────────────────────────────────────
def create_safety_compliance_pdf():
    path = str(BASE_DIR / "safety_compliance_report.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    story = []

    story.append(Paragraph("Workplace Safety & Environmental Compliance Report", title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph("1. Safety Performance Overview", heading_style))
    story.append(Paragraph(
        "SteelForge Manufacturing maintained a Total Recordable Incident Rate (TRIR) of 1.8 "
        "for Q1 2026, below the industry average of 3.2 for metal fabrication (NAICS 3323). "
        "There were zero lost-time injuries (LTI) during the quarter, extending the LTI-free "
        "streak to 247 consecutive days. Near-miss reporting increased 34% quarter-over-quarter "
        "following the launch of the anonymous digital reporting app, indicating improved safety "
        "culture and hazard awareness among shop floor personnel.",
        body_style
    ))
    story.append(Paragraph(
        "The Days Away, Restricted, or Transferred (DART) rate was 0.9, well below the company "
        "target of 1.5. All incidents were classified as first-aid only, primarily minor "
        "lacerations (3 events) and heat stress incidents (2 events) in the heat treatment "
        "department during the February ambient temperature spike.",
        body_style
    ))

    story.append(Paragraph("2. OSHA Compliance Status", heading_style))
    story.append(Paragraph(
        "The annual OSHA compliance audit was completed in March 2026 with zero citations. "
        "Key compliance areas reviewed included: Lockout/Tagout (LOTO) procedures per 29 CFR "
        "1910.147, Machine Guarding per 29 CFR 1910.212, Hazard Communication (HazCom) per "
        "29 CFR 1910.1200, Personal Protective Equipment (PPE) per 29 CFR 1910.132-138, and "
        "Confined Space Entry per 29 CFR 1910.146. All Safety Data Sheets (SDS) were verified "
        "current and accessible at each chemical use point.",
        body_style
    ))

    story.append(Paragraph("3. Environmental Compliance", heading_style))
    story.append(Paragraph(
        "Emissions monitoring confirmed compliance with EPA National Emission Standards for "
        "Hazardous Air Pollutants (NESHAP) for metal fabrication. VOC emissions from the "
        "painting and coating line averaged 4.2 tons/year against a permit limit of 10 tons/year. "
        "Wastewater discharge from the coolant treatment system met all NPDES permit parameters. "
        "Total Suspended Solids (TSS) averaged 12 mg/L against a limit of 30 mg/L. Oil and "
        "grease content averaged 8 mg/L against a limit of 15 mg/L.",
        body_style
    ))
    story.append(Paragraph(
        "Hazardous waste generation totaled 4.7 tons for Q1, classified under RCRA as D007 "
        "(chromium) from the chrome plating line and D008 (lead) from soldering operations. "
        "All waste was manifested and transported by licensed hauler EcoTransport LLC to the "
        "Clean Harbors facility in Deer Park, TX. The facility maintained Large Quantity "
        "Generator (LQG) status with all required biennial reports filed on time.",
        body_style
    ))

    data = [
        ["Parameter", "Permit Limit", "Q1 Average", "Status"],
        ["VOC Emissions", "10 tons/yr", "4.2 tons/yr", "COMPLIANT"],
        ["TSS Discharge", "30 mg/L", "12 mg/L", "COMPLIANT"],
        ["Oil & Grease", "15 mg/L", "8 mg/L", "COMPLIANT"],
        ["pH Range", "6.0 - 9.0", "7.2", "COMPLIANT"],
        ["Chromium (Cr)", "2.77 mg/L", "0.89 mg/L", "COMPLIANT"],
        ["Noise Level (boundary)", "70 dBA", "62 dBA", "COMPLIANT"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data))

    story.append(PageBreak())
    story.append(Paragraph("4. Training and Certification Status", heading_style))
    story.append(Paragraph(
        "All required safety training was completed on schedule. 100% of production employees "
        "completed annual refresher training in: Forklift Operation (OSHA 29 CFR 1910.178), "
        "Fire Extinguisher Use (29 CFR 1910.157), Emergency Action Plan (29 CFR 1910.38), and "
        "Bloodborne Pathogens (29 CFR 1910.1030). The new Robotic Cell Safety Training program "
        "was launched covering collaborative robot (cobot) interaction zones, emergency stop "
        "procedures, and risk assessment per ISO 10218 and ISO/TS 15066 standards. 87 operators "
        "have been certified to date, with the remaining 23 scheduled for April completion.",
        body_style
    ))

    story.append(Paragraph("5. Incident Investigation Summary", heading_style))
    data2 = [
        ["Date", "Type", "Department", "Severity", "Root Cause", "Status"],
        ["Jan 12", "Laceration", "Stamping", "First Aid", "Burr on part edge", "Closed"],
        ["Jan 28", "Laceration", "Assembly", "First Aid", "Sharp fixture edge", "Closed"],
        ["Feb 14", "Heat stress", "Heat Treat", "First Aid", "HVAC malfunction", "Closed"],
        ["Feb 15", "Heat stress", "Heat Treat", "First Aid", "HVAC malfunction", "Closed"],
        ["Mar 03", "Laceration", "CNC Shop", "First Aid", "Chip ejection", "Closed"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data2))

    doc.build(story)
    print(f"  Created: {path}")


# ── PDF 4: Maintenance & Reliability Report ────────────────────────────────
def create_maintenance_pdf():
    path = str(BASE_DIR / "maintenance_reliability_report.pdf")
    doc = SimpleDocTemplate(path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    story = []

    story.append(Paragraph("Maintenance & Reliability Engineering Report — Q1 2026", title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph("1. Maintenance Strategy Overview", heading_style))
    story.append(Paragraph(
        "SteelForge Manufacturing employs a three-tier maintenance strategy: Preventive Maintenance "
        "(PM), Predictive Maintenance (PdM), and Reliability-Centered Maintenance (RCM). The "
        "Computerized Maintenance Management System (CMMS) — SAP Plant Maintenance module — tracks "
        "all work orders, PM schedules, spare parts inventory, and equipment history. Total "
        "maintenance spend for Q1 was $1.87M against a budget of $2.1M (89% budget utilization).",
        body_style
    ))
    story.append(Paragraph(
        "The ratio of planned to unplanned maintenance improved from 72:28 in Q4 2025 to 81:19 "
        "in Q1 2026, exceeding the world-class benchmark of 80:20. Mean Time Between Failures "
        "(MTBF) for critical equipment increased 18% to 1,247 hours. Mean Time To Repair (MTTR) "
        "decreased 12% to 2.3 hours, reflecting investments in technician training and spare "
        "parts availability improvements.",
        body_style
    ))

    story.append(Paragraph("2. Predictive Maintenance Program", heading_style))
    story.append(Paragraph(
        "The PdM program covers vibration analysis, infrared thermography, oil analysis, and "
        "ultrasonic testing across 156 critical assets. Key findings from Q1: Vibration analysis "
        "detected early-stage bearing degradation on CNC Line 4 spindle assembly (Asset #CNC-A-004) "
        "in January, but the replacement was deferred to the February maintenance window. The "
        "bearing failed catastrophically on Feb 8, causing the 36-hour unplanned shutdown. This "
        "event prompted a revision of the PdM decision matrix to lower the deferral threshold "
        "for spindle-class bearings from ISO 10816 Zone C to Zone B.",
        body_style
    ))

    data = [
        ["Asset ID", "Equipment", "PdM Method", "Finding", "Action Taken"],
        ["CNC-A-004", "Mori Seiki NHX5000", "Vibration", "Bearing ISO Zone C→D", "Replaced (after failure)"],
        ["STAMP-A-002", "Komatsu H2F600", "Thermography", "Motor hot spot 142°C", "Rewound motor — PM"],
        ["WELD-B-007", "Fanuc ArcMate 120", "Oil analysis", "Fe >150ppm", "Gearbox rebuild"],
        ["HT-B-001", "Ipsen TurboTreater", "Ultrasonic", "Wall thinning 0.8mm", "Scheduled Q2 repair"],
        ["COAT-B-003", "Nordson powder booth", "Thermography", "Oven element failing", "Replaced element"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data, col_widths=[0.9*inch, 1.3*inch, 0.9*inch, 1.5*inch, 1.5*inch]))

    story.append(Paragraph("3. Spare Parts and Inventory Management", heading_style))
    story.append(Paragraph(
        "The spare parts storeroom holds $4.2M in inventory across 8,700 SKUs. Inventory turns "
        "improved from 1.8x to 2.1x through the implementation of min/max reorder points based "
        "on equipment criticality ranking and historical consumption. Obsolete inventory write-off "
        "for Q1 was $127K, primarily related to legacy Mazak control boards replaced during the "
        "CNC retrofit program. A consignment agreement was established with SKF for bearing stock, "
        "reducing on-hand inventory by $310K while maintaining 98% service level.",
        body_style
    ))

    story.append(Paragraph("4. Reliability Metrics Dashboard", heading_style))
    data2 = [
        ["Metric", "Q4 2025", "Q1 2026", "Target", "Trend"],
        ["MTBF (hours)", "1,056", "1,247", "1,300", "↑ Improving"],
        ["MTTR (hours)", "2.6", "2.3", "2.0", "↑ Improving"],
        ["PM Compliance", "88%", "94%", "95%", "↑ Improving"],
        ["Planned:Unplanned", "72:28", "81:19", "80:20", "✓ Target Met"],
        ["Spare Parts Service Level", "94%", "98%", "97%", "✓ Target Met"],
        ["Maintenance Cost/Unit", "$1.14", "$1.01", "$1.00", "↑ Improving"],
    ]
    story.append(Spacer(1, 10))
    story.append(build_table(data2))

    doc.build(story)
    print(f"  Created: {path}")


# ── TXT 1: Standard Operating Procedure ────────────────────────────────────
def create_sop_txt():
    path = BASE_DIR / "sop_cnc_machining.txt"
    path.write_text("""\
STANDARD OPERATING PROCEDURE: CNC MACHINING CENTER OPERATION
Document ID: SOP-CNC-001 | Revision: 7 | Effective Date: 2026-01-15
Department: CNC Machining | Plant: Plant A
Prepared by: James Rodriguez, Manufacturing Engineer
Approved by: Sarah Chen, Quality Manager

=== 1. PURPOSE ===
This SOP defines the step-by-step procedure for operating CNC machining centers
(Mori Seiki NHX5000 and DMG Mori NLX2500) at SteelForge Manufacturing Plant A.
It ensures consistent part quality, operator safety, and equipment longevity.

=== 2. SCOPE ===
Applies to all CNC operators (Grade C3 and above), setup technicians, and
production supervisors in the CNC Machining department. Covers operations for
aluminum, steel, stainless steel, and titanium workpieces.

=== 3. SAFETY REQUIREMENTS ===
- Mandatory PPE: Safety glasses (ANSI Z87.1), steel-toe boots, hearing
  protection (when noise > 85 dBA), cut-resistant gloves (ANSI A4 rating)
- Verify machine guarding is intact and interlocks functional before startup
- Confirm emergency stop (E-stop) buttons are accessible and operational
- Never reach into the work envelope while spindle is rotating
- Ensure coolant concentration is between 6-8% (check with refractometer)
- Lock out/Tag out per SOP-LOTO-001 before any tool change or adjustment

=== 4. PRE-OPERATION CHECKLIST ===
4.1 Machine Warm-Up
    - Power on machine and allow 15-minute warm-up cycle
    - Run spindle warm-up program (O9001) at 500, 2000, 5000, 8000 RPM
    - Verify axis home positions using reference marks
    - Check hydraulic pressure: 70-80 bar nominal
    - Verify coolant level in sump (minimum 75% full)
    - Check chip conveyor operation

4.2 Tooling Setup
    - Load tool list from job sheet (MES terminal)
    - Verify tool offsets using Renishaw OMP60 probe
    - Confirm tool life remaining in Tool Management System (TMS)
    - Replace any tools below 20% remaining life
    - Verify collet/chuck runout < 0.005mm using dial indicator

4.3 Work Holding
    - Select fixture per engineering drawing callout
    - Verify fixture locating surfaces are clean and free of chips
    - Torque fixture bolts per specification (typically 45 Nm for M12)
    - Verify part seating using air-sensing circuit (green light = seated)

=== 5. PRODUCTION OPERATION ===
5.1 First Article Run
    - Run first piece at 80% programmed feed rate
    - Perform 100% dimensional inspection per FAI checklist
    - Record measurements in MES quality module
    - Obtain Quality Inspector sign-off before production release
    - Retain first article as reference sample in labeled bag

5.2 Normal Production
    - Run program at programmed speeds and feeds
    - Monitor tool wear indicators on HMI display
    - Perform in-process gauging per control plan frequency:
      * Critical dimensions: every 5th part
      * Standard dimensions: every 25th part
      * Surface finish (Ra): every 50th part
    - Record SPC data points in real-time via shop floor terminal
    - Flag any Cpk < 1.33 to Quality Engineer immediately

5.3 Tool Change Procedure
    - When TMS signals tool life expiry, pause program at safe point
    - Follow LOTO abbreviated procedure (SOP-LOTO-001 Section 4.2)
    - Replace tool and reset tool offset using probe cycle
    - Run single-part verification at reduced feed rate
    - Resume production after dimensional confirmation

=== 6. POST-OPERATION ===
    - Run machine clean-up cycle (O9099)
    - Remove all workpieces and tooling from table
    - Clean chip pan and empty chip hopper
    - Top off coolant if below 80% level
    - Log production quantity and downtime in MES
    - Complete shift handover form (Form SH-201)
    - Report any machine anomalies on CMMS work request

=== 7. TROUBLESHOOTING GUIDE ===
| Symptom                    | Probable Cause            | Action                    |
|---------------------------|--------------------------|---------------------------|
| Chatter marks on surface  | Tool wear / wrong RPM    | Replace tool, verify RPM  |
| Dimension out of tolerance| Thermal drift / tool wear | Re-probe, check coolant   |
| Poor surface finish (Ra)  | Feed rate too high        | Reduce feed 10-15%        |
| Coolant foaming           | Concentration too high    | Dilute to 6-8% range      |
| Spindle alarm             | Overload / bearing issue  | Stop, call maintenance    |
| Chip evacuation problem   | Conveyor jam              | Clear jam, check auger    |

=== 8. REVISION HISTORY ===
Rev 7 (2026-01-15): Added titanium machining parameters, updated PdM references
Rev 6 (2025-07-01): Updated tool life management per TMS upgrade
Rev 5 (2025-01-10): Added cobot loading integration procedure
Rev 4 (2024-06-15): Revised safety section per new OSHA guidance
""")
    print(f"  Created: {path}")


# ── TXT 2: Supply Chain Procedures ─────────────────────────────────────────
def create_supply_chain_txt():
    path = BASE_DIR / "supply_chain_procedures.txt"
    path.write_text("""\
SUPPLY CHAIN MANAGEMENT PROCEDURES
Document ID: SCM-PROC-001 | Revision: 4 | Effective Date: 2026-02-01
Department: Procurement & Supply Chain | SteelForge Manufacturing
Prepared by: Maria Santos, Supply Chain Director

=== 1. PROCUREMENT PROCESS ===

1.1 Purchase Requisition (PR)
    All material purchases must originate from an approved Purchase Requisition
    in the ERP system (SAP MM module). PRs are auto-generated by MRP for
    production materials based on BOM explosion and demand forecasting. Non-production
    purchases require department manager approval for amounts under $5,000 and
    VP-level approval for amounts exceeding $5,000.

1.2 Supplier Selection
    New suppliers must complete the Supplier Qualification Questionnaire (Form SQ-100)
    and pass an on-site audit scoring minimum 80/100. Evaluation criteria:
    - Quality Management System (ISO 9001 minimum, IATF 16949 preferred): 30 pts
    - Delivery Performance history: 20 pts
    - Financial Stability (D&B rating): 15 pts
    - Technical Capability assessment: 20 pts
    - Cost Competitiveness: 15 pts

1.3 Approved Vendor List (AVL)
    Only suppliers on the AVL may receive purchase orders. The AVL is reviewed
    quarterly by the Commodity Management team. Key suppliers by commodity:

    Steel (flat-rolled):
    - Primary: Nippon Steel Corp — Lead time 8-12 weeks, MOQ 20 tons
    - Secondary: ArcelorMittal — Lead time 6-10 weeks, MOQ 15 tons
    - Emergency: Steel Warehouse Inc — Lead time 1-2 weeks, premium +22%

    Aluminum:
    - Primary: Alcoa Corporation — Lead time 6-8 weeks
    - Secondary: Novelis Inc — Lead time 8-10 weeks

    Cutting Tools:
    - Sandvik Coromant — Consignment program, 48-hour replenishment
    - Kennametal — Standard PO, 2-week lead time
    - Iscar — Secondary source for specialty tooling

=== 2. INVENTORY MANAGEMENT ===

2.1 Classification
    Inventory is managed using ABC-XYZ classification:
    - A items (top 80% of annual spend): Weekly review, safety stock = 3 weeks
    - B items (next 15% of spend): Bi-weekly review, safety stock = 4 weeks
    - C items (bottom 5% of spend): Monthly review, safety stock = 6 weeks
    - X (stable demand): Standard reorder point
    - Y (variable demand): Dynamic safety stock with demand sensing
    - Z (sporadic demand): Order-to-demand only

2.2 Warehouse Operations
    Raw material warehouse operates 24/5 with 2 receiving docks and 3 shipping
    docks. Inbound materials are received through GR (Goods Receipt) process
    with barcode scanning for lot traceability. Storage follows FIFO (First In
    First Out) methodology enforced through WMS (Warehouse Management System)
    directed putaway and picking. Annual physical inventory count is conducted
    in December; cycle counts for A-items occur weekly.

2.3 Kanban System
    Shop floor consumables and fasteners use a 2-bin kanban system. Each bin
    is sized for 1 week of consumption. When a bin empties, the kanban card
    triggers automatic replenishment via the WMS. Supplier-managed inventory
    (SMI) is used for high-volume fasteners with weekly vendor visits.

=== 3. LOGISTICS AND SHIPPING ===

3.1 Inbound Logistics
    - Domestic steel shipments: Flatbed truck, 2-5 day transit
    - Import materials: Ocean freight via Port of Houston, 6-8 week transit
    - Expedited: Air freight authorized only by Supply Chain Director for
      critical production shortages (cost > 10x standard shipping)

3.2 Outbound Logistics
    - Standard: LTL or FTL via contract carriers (XPO, Old Dominion)
    - Customer-specific: Some OEM customers require dedicated carriers
    - International: Freight forwarder coordination for customs/duties
    - All shipments require ASN (Advance Shipping Notice) transmitted via EDI 856
    - Packaging per AIAG guidelines or customer-specific requirements

=== 4. SUPPLY CHAIN RISK MANAGEMENT ===

4.1 Risk Assessment
    Quarterly risk assessments are conducted for all Tier-1 suppliers using a
    Risk Priority Number (RPN) methodology. Factors include:
    - Geographic concentration risk (single-source, single-country)
    - Financial health monitoring via D&B alerts
    - Natural disaster exposure (earthquake, flood, hurricane zones)
    - Geopolitical risk for international suppliers
    - Cybersecurity posture assessment

4.2 Business Continuity
    Critical materials (steel, aluminum, titanium) must have minimum 2 qualified
    sources. Single-source items require 6-week safety stock buffer. The Business
    Continuity Plan (BCP) is tested annually through tabletop exercises simulating
    supplier disruption scenarios.

=== 5. KEY METRICS ===
    Metric                          Target      Q1 2026 Actual
    ----------------------------------------------------------
    On-Time Delivery (inbound)      95%         93.7%
    Purchase Price Variance         < 2%        1.8%
    Inventory Turns (raw material)  12x/year    10.4x
    Supplier Quality (incoming PPM) < 500       387
    Emergency Purchases (% of POs)  < 3%        2.1%
    Perfect Order Rate              > 92%       94.2%
""")
    print(f"  Created: {path}")


# ── Excel: Production Metrics Dashboard ────────────────────────────────────
def create_production_metrics_excel():
    path = str(BASE_DIR / "production_metrics_q1_2026.xlsx")
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    num_fmt = '#,##0'
    pct_fmt = '0.0%'

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data(ws, start_row, end_row):
        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Daily Production Output ──
    ws1 = wb.active
    ws1.title = "Daily Production"
    ws1.append(["Date", "Line", "Shift", "Part Number", "Planned Qty",
                "Actual Qty", "Scrap Qty", "Scrap %", "OEE %", "Downtime (min)"])
    style_header(ws1)

    import random
    random.seed(42)
    lines = ["CNC-A-001", "CNC-A-002", "CNC-A-003", "STAMP-A-001", "STAMP-A-002",
             "WELD-A-001", "WELD-B-001", "HT-B-001"]
    parts = ["TH-4401", "TH-4402", "BRK-2200", "BRK-2201", "SFT-3300",
             "HSG-1100", "HSG-1101", "GR-5500"]
    shifts = ["Day", "Evening", "Night"]

    row_num = 2
    for month in range(1, 4):
        for day in range(1, 29):
            for _ in range(random.randint(2, 4)):
                line = random.choice(lines)
                part = random.choice(parts)
                shift = random.choice(shifts)
                planned = random.randint(200, 800)
                scrap = random.randint(0, int(planned * 0.04))
                actual = planned - random.randint(0, int(planned * 0.08)) - scrap
                oee = round(random.uniform(0.82, 0.99), 3)
                downtime = random.randint(0, 90)
                date_str = f"2026-{month:02d}-{day:02d}"
                ws1.append([date_str, line, shift, part, planned, actual, scrap,
                           scrap / planned if planned else 0, oee, downtime])
                ws1.cell(row=row_num, column=8).number_format = pct_fmt
                ws1.cell(row=row_num, column=9).number_format = pct_fmt
                row_num += 1

    style_data(ws1, 2, row_num - 1)
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 14

    # ── Sheet 2: Equipment Downtime Log ──
    ws2 = wb.create_sheet("Downtime Log")
    ws2.append(["Date", "Asset ID", "Equipment Name", "Downtime (min)",
                "Category", "Root Cause", "Shift", "Technician"])
    style_header(ws2)

    categories = ["Mechanical", "Electrical", "Tooling", "Quality", "Material",
                  "Changeover", "Planned PM", "Operator"]
    causes = [
        "Bearing failure", "Drive belt worn", "Tool breakage", "Dimension drift",
        "Material delay", "Die changeover", "Scheduled PM", "Training",
        "Hydraulic leak", "PLC fault", "Probe calibration", "Coolant issue",
        "Sensor malfunction", "Power fluctuation", "Fixture damage", "Program error"
    ]
    technicians = ["Mike T.", "Lisa K.", "David R.", "Priya M.", "Carlos G.", "Anna W."]
    equipment = {
        "CNC-A-001": "Mori Seiki NHX5000 #1", "CNC-A-002": "Mori Seiki NHX5000 #2",
        "CNC-A-003": "DMG Mori NLX2500", "STAMP-A-001": "Komatsu H2F600 #1",
        "STAMP-A-002": "Komatsu H2F300", "WELD-A-001": "Fanuc ArcMate 120iC",
        "WELD-B-001": "Fanuc ArcMate 120iC #2", "HT-B-001": "Ipsen TurboTreater",
    }

    row_num = 2
    for month in range(1, 4):
        for day in range(1, 29):
            for _ in range(random.randint(0, 3)):
                asset = random.choice(list(equipment.keys()))
                ws2.append([
                    f"2026-{month:02d}-{day:02d}", asset, equipment[asset],
                    random.randint(5, 180), random.choice(categories),
                    random.choice(causes), random.choice(shifts),
                    random.choice(technicians)
                ])
                row_num += 1

    style_data(ws2, 2, row_num - 1)
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20

    # ── Sheet 3: Supplier Scorecard ──
    ws3 = wb.create_sheet("Supplier Scorecard")
    ws3.append(["Supplier", "Commodity", "Quality Score", "Delivery Score",
                "Cost Score", "Overall Score", "Status", "PPM", "On-Time %"])
    style_header(ws3)

    suppliers = [
        ("Nippon Steel Corp", "Steel", 92, 88, 85, 89, "Approved", 210, 0.94),
        ("ArcelorMittal", "Steel", 88, 91, 90, 89, "Approved", 340, 0.96),
        ("Steel Warehouse Inc", "Steel", 78, 95, 65, 78, "Conditional", 580, 0.98),
        ("Alcoa Corporation", "Aluminum", 95, 90, 82, 90, "Preferred", 120, 0.95),
        ("Novelis Inc", "Aluminum", 87, 85, 88, 86, "Approved", 290, 0.91),
        ("Sandvik Coromant", "Cutting Tools", 97, 96, 78, 92, "Preferred", 45, 0.99),
        ("Kennametal", "Cutting Tools", 91, 88, 85, 88, "Approved", 150, 0.93),
        ("Iscar", "Cutting Tools", 85, 82, 90, 85, "Approved", 220, 0.89),
        ("SKF", "Bearings", 94, 92, 80, 90, "Preferred", 80, 0.97),
        ("Timken", "Bearings", 90, 87, 83, 87, "Approved", 160, 0.94),
        ("Parker Hannifin", "Hydraulics", 93, 89, 81, 88, "Approved", 95, 0.95),
        ("Festo", "Pneumatics", 96, 94, 77, 91, "Preferred", 55, 0.98),
    ]
    for i, s in enumerate(suppliers, start=2):
        ws3.append(list(s))
        ws3.cell(row=i, column=9).number_format = pct_fmt

    style_data(ws3, 2, len(suppliers) + 1)
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 18

    # ── Sheet 4: Cost Analysis ──
    ws4 = wb.create_sheet("Cost Analysis")
    ws4.append(["Month", "Category", "Budget ($)", "Actual ($)", "Variance ($)",
                "Variance %", "Notes"])
    style_header(ws4)

    cost_data = [
        ("Jan", "Raw Materials", 2400000, 2520000, -120000, -0.05, "Steel price increase"),
        ("Jan", "Direct Labor", 1800000, 1785000, 15000, 0.008, "On target"),
        ("Jan", "Maintenance", 700000, 645000, 55000, 0.079, "Under budget - deferred PM"),
        ("Jan", "Utilities", 320000, 348000, -28000, -0.088, "Winter heating costs"),
        ("Jan", "Tooling", 180000, 192000, -12000, -0.067, "Tool breakage on Ti parts"),
        ("Feb", "Raw Materials", 2400000, 2380000, 20000, 0.008, "On target"),
        ("Feb", "Direct Labor", 1800000, 1920000, -120000, -0.067, "OT for CNC-004 recovery"),
        ("Feb", "Maintenance", 700000, 890000, -190000, -0.271, "CNC-004 spindle replacement"),
        ("Feb", "Utilities", 320000, 310000, 10000, 0.031, "Mild weather"),
        ("Feb", "Tooling", 180000, 175000, 5000, 0.028, "On target"),
        ("Mar", "Raw Materials", 2500000, 2460000, 40000, 0.016, "Volume ramp TH-4400"),
        ("Mar", "Direct Labor", 1850000, 1840000, 10000, 0.005, "On target"),
        ("Mar", "Maintenance", 700000, 680000, 20000, 0.029, "On target"),
        ("Mar", "Utilities", 300000, 295000, 5000, 0.017, "On target"),
        ("Mar", "Tooling", 200000, 215000, -15000, -0.075, "New TH-4400 tooling setup"),
    ]
    for i, row in enumerate(cost_data, start=2):
        ws4.append(list(row))
        ws4.cell(row=i, column=3).number_format = num_fmt
        ws4.cell(row=i, column=4).number_format = num_fmt
        ws4.cell(row=i, column=5).number_format = '#,##0'
        ws4.cell(row=i, column=6).number_format = pct_fmt

    style_data(ws4, 2, len(cost_data) + 1)
    for col in ws4.columns:
        ws4.column_dimensions[col[0].column_letter].width = 18

    wb.save(path)
    print(f"  Created: {path}")


if __name__ == "__main__":
    print("Creating sample manufacturing documents...")
    create_quality_control_pdf()
    create_production_planning_pdf()
    create_safety_compliance_pdf()
    create_maintenance_pdf()
    create_sop_txt()
    create_supply_chain_txt()
    create_production_metrics_excel()
    print("\nAll sample documents created successfully!")
